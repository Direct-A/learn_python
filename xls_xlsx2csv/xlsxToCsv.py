#coding: utf-8
# 依赖openpyxl库：http://openpyxl.readthedocs.org/en/latest/
import csv
import os
import sys
from os.path import splitext

import xlrd
import openpyxl
from tqdm import tqdm


def xlsx2csv(filename):
    try:
        xlsx_file = openpyxl.load_workbook(filename=filename)

        # 统计总工作量
        sum_cell = {}
        sum_cell_num = 0
        for sheet in xlsx_file.sheetnames:
            col = xlsx_file[sheet].max_column
            row = xlsx_file[sheet].max_row
            cell = col * row
            sum_cell_num += cell
            cell_num = {sheet: cell}
            sum_cell.update(cell_num)

        with tqdm(total=sum_cell_num) as pbar:
            pbar.set_description('Processing')

            pros_count = 0
            for sheet in xlsx_file.sheetnames:
                pros_count += sum_cell[sheet]
                # 每个sheet输出到一个csv文件中，文件名用xlsx文件名和sheet名用'_'连接
                csv_filename = '{xlsx}_{sheet}.csv'.format(
                    xlsx=os.path.splitext(filename.replace(' ', '-'))[0],
                    sheet=sheet.replace(' ', '-'))

                # creat files and open as binary
                try:
                    csv_file = open(csv_filename, 'x')
                except FileExistsError:
                    csv_file = open(csv_filename, 'w')
                csv_file_writer = csv.writer(csv_file, delimiter='\t')

                # write xls cells into csv
                sheet_ranges = xlsx_file[sheet]
                for row in sheet_ranges.rows:
                    row_container = []
                    for cell in row:
                        if type(cell.value) == 'unicode':
                            row_container.append(cell.value.encode('utf-8'))
                        else:
                            row_container.append(str(cell.value))
                    csv_file_writer.writerow(row_container)
                csv_file.close()
                pbar.update(pros_count)

    except Exception as e:
        print(e)


def xls2csv(filename):
    try:
        xls_file = xlrd.open_workbook(filename=filename)

        # 进度条：统计总工作量
        sum_cell = {}
        sum_cell_num = 0
        for sheet in xls_file.sheet_names():
            col = xls_file.sheet_by_name(sheet).nrows
            row = xls_file.sheet_by_name(sheet).ncols
            cell = col * row
            sum_cell_num += cell
            cell_num = {sheet: cell}
            sum_cell.update(cell_num)

        with tqdm(total=sum_cell_num) as pbar:
            pbar.set_description('Processing')
            pros_count = 0

            for sheet in xls_file.sheet_names():
                # 单sheet进度条增加量
                pros_count += sum_cell[sheet]

                # 每个sheet输出到一个csv文件中，文件名用xls文件名和sheet名用'_'连接
                csv_filename = '{xlsx}_{sheet}.csv'.format(
                    xls=os.path.splitext(filename.replace(' ', '-'))[0],
                    sheet=sheet.replace(' ', '-'))

                # creat files and open as binary
                try:
                    csv_file = open(csv_filename, 'x')
                except FileExistsError:
                    csv_file = open(csv_filename, 'w')
                csv_file_writer = csv.writer(csv_file, delimiter='\t')

                # write xls cells into csv
                sheet_ranges_row = range(0, xls_file.sheet_by_name(sheet).nrows)
                sheet_ranges_col = range(0, xls_file.sheet_by_name(sheet).ncols)
                for rowx in sheet_ranges_row:
                    row_container = []
                    for colx in sheet_ranges_col:
                        cell = xls_file.sheet_by_name(sheet).cell_value(rowx, colx)
                        if type(cell.value) == 'unicode':
                            row_container.append(cell.value.encode('utf-8'))
                        else:
                            row_container.append(str(cell.value))
                    csv_file_writer.writerow(row_container)
                csv_file.close()

                # 更新进度条
                pbar.update(pros_count)

    except Exception as e:
        print(e)


#################################
# execute only if run as a script
#################################
if __name__ == '__main__':
    # if len(sys.argv) != 2:
    #    print('usage: xlsx2csv <xlsx file name>')
    # else:
    #    filename = input("Give me FULL PATH of xlsx_files\nfor example: ~/#Download/test.xlsx\n>>")
    #    xlsx2csv(sys.argv[1])
    while True:
        filename = input(
            "\033[1;32;40mGive me \033[1;31;43mFULL PATH\033[1;32;40m of xlsx_files\n\033[0;32;40mFor Example: ~/Download/test.xlsx\033[0m\n>>")

        if filename in ["q", "Q", "e", "exit"]:
            sys.exit(0)
        else:
            ext = os.path.splitext(filename)

            if ext[1] == ".xls":
                xls2csv(filename)
            elif ext[1] == ".xlsx":
                xlsx2csv(filename)
