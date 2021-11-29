#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Author: sYc
# Date: 2021-03-26 20:39:53
# Version: 0.1
# Description: douban movie top 250

import openpyxl
import requests
from bs4 import BeautifulSoup


def request_db(url: str) -> object:
    head = {
        "user-agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1"
    }
    data_all = requests.get(url, headers=head)
    if data_all.status_code == 200:
        soup = BeautifulSoup(data_all.text, "html.parser")
        return soup
    else:
        return print("Error: {}".format(data_all.status_code))


def tidier(ResultSet, garbs: list):
    for garb in garbs:
        ResultSet = ResultSet.replace(str(garb), "")
    return ResultSet


def cutter(inner_soup: object, selector: str, flag: int = 1, **sp) -> str:
    ResultSet = inner_soup.select(selector)
    if ResultSet == []:
        ResultSet = BeautifulSoup("", "html.parser")
    else:
        ResultSet = ResultSet[0]

    def inner_get_text():
        result = ResultSet.get_text(strip=True)
        return tidier(result, [" "])

    def inner_get(sp: str):
        result = ResultSet.get(sp)
        return tidier(result, [" "])

    chooser = {
        1: inner_get_text,
        2: inner_get
    }

    return chooser[flag](**sp)


def main(page: int):
    url = "https://movie.douban.com/top250?start={}&filter=".format(page*25)
    soup = request_db(url)
    data = soup.select("#content > div > div.article > ol > li")

    for block in data:
        inner_soup = BeautifulSoup(str(block), "html.parser")
        name = cutter(inner_soup, "div.hd > a")
        pic = cutter(inner_soup, "div.pic > a > img", 2, sp="src")
        rank = cutter(inner_soup, "div.pic > em")
        rating = cutter(
            inner_soup, "div.info > div.bd > div > span.rating_num")
        director = cutter(
            inner_soup, "div.info > div.bd > p:nth-child(1)")
        intro = cutter(
            inner_soup, "div.info > div.bd > p.quote > span")
        data_block = {
            "name": name,
            "pic": pic,
            "rank": rank,
            "rating": rating,
            "director": director,
            "intro": intro,
        }

        data_continner.append(data_block)


def write_to_file():
    for row, data_block in zip(range(2, len(data_continner)+2), data_continner):
        for col, key in zip(c_range, keys):
            ws.cell(row, col, data_block[key])

    wb.save(file_name)


if __name__ == "__main__":
    file_name = "data_db_movietop250.xlsx"

    # 判断文件是否创建
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.save("db_movie_top250.xlsx")
        ws = wb.active

    # 写入表头
    keys = ["name", "pic", "rank", "rating", "director", "intro"]
    c_range = range(1, 7)
    for col, key in zip(c_range, keys):
        ws.cell(1, col, key)
    wb.save(file_name)

    data_continner = []

    for page in range(20):
        main(page)

    write_to_file()
