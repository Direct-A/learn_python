# from chromedriver_binary import add_chromedriver_to_path
# from selenium import webdriver

# add_chromedriver_to_path()
# proxy="127.0.0.1:7890"
# chrome_options = webdriver.ChromeOptions()
# # chrome_options.add_argument("--headless")#设置无头模式
# chrome_options.add_argument('--proxy-server={}'.format(proxy))#设置代理
# driver = webdriver.Chrome(chrome_options=chrome_options)
# driver.get("https://www.baidu.com")

import openpyxl
from click.formatting import iter_rows

wb = openpyxl.Workbook()
ws = wb.active
ws["g1"] = 25

for cell, i in ws[1], range(7):
    j = cell
    j.value = i

for row in ws.iter_rows(values_only=True):
    for cell in row:
        print(cell)
